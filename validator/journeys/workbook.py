"""
workbook.py — Multi-sheet Excel export of the full agent analysis.

The Dialogflow CX equivalent of the 9-sheet workbook described in
VOICEBOT_INTENT_ANALYSIS_PLAYBOOK.md §4, built from data the validator has
already computed:

  1. Overview               Health and verdict tallies, check summary, themes
  2. Journey Analysis       One row per head intent — shape plus narrative
  3. Response vs Scope      Purpose beside the actual bot responses
  4. Coverage Gaps          Unrouted intents, no-agent-path journeys, RG gaps
  5. Overlap & Confusion    Exact and near-duplicate intent collisions
  6. ASR & Noise Phrases    Known mistranscriptions kept as training data
  7. Data Hygiene           Empty phrases, duplicates, metadata drift
  8. Phrase Counts          Per-intent volume with health rating
  9. Validation Findings    Every finding from every check

openpyxl is an optional dependency. If it is not installed, is_available()
returns False and validate.py skips the workbook rather than failing — the
tool stays usable in a stdlib-only CI environment.
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from validator.loader import AgentIndex
from validator.checks.models import Finding
from validator.checks.nlu import (
    exact_collisions, near_duplicate_pairs, asr_hits, phrase_stats,
)
from validator.journeys.tracer import Journey

try:  # pragma: no cover - import guard
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    _AVAILABLE = True
except ImportError:  # pragma: no cover - import guard
    _AVAILABLE = False


def is_available() -> bool:
    """True if openpyxl is installed and the workbook can be written."""
    return _AVAILABLE


# ── Styling ───────────────────────────────────────────────────────────────────

_HEADER_FILL = "1F3864"
_HEADER_FONT = "FFFFFF"

# Colour-coding shared by health, verdict and severity columns.
_RED = "F8CBAD"
_AMBER = "FFE699"
_GREEN = "C6E0B4"
_GREY = "E7E6E6"

_VALUE_COLOURS = {
    # Health
    "major issue": _RED,
    "needs attention": _AMBER,
    "good": _GREEN,
    # Response verdict
    "missing": _RED,
    "misaligned": _RED,
    "partial": _AMBER,
    "no scope": _AMBER,
    "accurate": _GREEN,
    "by design": _GREY,
    # Severity
    "error": _RED,
    "warning": _AMBER,
    "pass": _GREEN,
    # Coverage / collision types
    "unrouted": _RED,
    "exact": _RED,
    "near-duplicate": _AMBER,
}

# Columns whose values drive cell colouring.
_COLOURED_HEADERS = {
    "Health", "Verdict", "Response verdict", "Severity", "Type", "Status",
}

# Sort keys that put the worst rows first, matching the playbook.
_HEALTH_ORDER = {"Major issue": 0, "Needs attention": 1, "Good": 2, "": 3}
_VERDICT_ORDER = {
    "MISSING": 0, "MISALIGNED": 1, "PARTIAL": 2, "NO SCOPE": 3,
    "ACCURATE": 4, "BY DESIGN": 5, "": 6,
}


def write_workbook(
    path: Path,
    agent: AgentIndex,
    journeys: list[Journey],
    narratives: dict[str, dict[str, str]],
    results: list[tuple[str, list[Finding]]],
) -> None:
    """
    Build and write the analysis workbook.

    Args:
        path:       Destination .xlsx path.
        agent:      Loaded AgentIndex.
        journeys:   Journey records from tracer.trace_journeys().
        narratives: Narrative analysis keyed by intent name.
        results:    (check_label, findings) pairs from the check run.

    Raises:
        RuntimeError: If openpyxl is not installed. Call is_available() first.
    """
    if not _AVAILABLE:
        raise RuntimeError(
            "openpyxl is not installed — run 'pip install openpyxl' to enable "
            "the Excel workbook, or omit --excel."
        )

    wb = Workbook()
    wb.remove(wb.active)  # Drop the default sheet; every sheet is added below.

    _sheet_overview(wb, agent, journeys, narratives, results)
    _sheet_journey_analysis(wb, journeys, narratives)
    _sheet_response_vs_scope(wb, journeys, narratives)
    _sheet_coverage_gaps(wb, agent, journeys, results)
    _sheet_overlap(wb, agent)
    _sheet_asr(wb, agent)
    _sheet_data_hygiene(wb, agent)
    _sheet_phrase_counts(wb, agent)
    _sheet_findings(wb, results)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_overview(
    wb: "Workbook",
    agent: AgentIndex,
    journeys: list[Journey],
    narratives: dict[str, dict[str, str]],
    results: list[tuple[str, list[Finding]]],
) -> None:
    ws = wb.create_sheet("Overview")

    analysed = [j for j in journeys if narratives.get(j.intent)]
    health = Counter(
        narratives[j.intent].get("health", "") for j in analysed
    )
    verdict = Counter(
        narratives[j.intent].get("response_verdict", "") for j in analysed
    )
    shapes = Counter(_journey_shape(j) for j in journeys)

    rows: list[list[Any]] = []

    rows.append(["AGENT", "", ""])
    rows.append(["Agent", agent.root.resolve().name, ""])
    rows.append(["Flows", len(agent.flow_files), ""])
    rows.append(["Pages", len(agent.page_files), ""])
    rows.append(["Intents", len(agent.intents), ""])
    rows.append(["Head intents", len(agent.head_intents), ""])
    rows.append(["Entity types", len(agent.entity_types), ""])
    rows.append(["Route groups", len(agent.route_groups), ""])
    rows.append(["Test cases", len(agent.test_cases), ""])
    rows.append(["", "", ""])

    rows.append(["JOURNEY SHAPE", "Count", "Share of head intents"])
    for shape, count in shapes.most_common():
        rows.append([shape, count, _pct(count, len(journeys))])
    rows.append(["", "", ""])

    rows.append(["NARRATIVE COVERAGE", "Count", ""])
    rows.append(["Journeys analysed", len(analysed), _pct(len(analysed), len(journeys))])
    rows.append(["Not yet analysed", len(journeys) - len(analysed), ""])
    rows.append(["", "", ""])

    rows.append(["JOURNEY HEALTH", "Count", "Share of analysed"])
    for label in ("Major issue", "Needs attention", "Good"):
        if health.get(label):
            rows.append([label, health[label], _pct(health[label], len(analysed))])
    rows.append(["", "", ""])

    rows.append(["RESPONSE VERDICT", "Count", "Share of analysed"])
    for label, count in verdict.most_common():
        if label:
            rows.append([label, count, _pct(count, len(analysed))])
    rows.append(["", "", ""])

    rows.append(["VALIDATION CHECKS", "Errors", "Warnings"])
    for label, findings in results:
        errors = sum(1 for f in findings if f.severity == "error")
        warnings = sum(1 for f in findings if f.severity == "warning")
        rows.append([label, errors, warnings])

    _write(ws, ["Metric", "Value", "Detail"], rows, widths=[38, 16, 22])

    # Bold the section banner rows so the sheet reads as blocks.
    for row_idx, row in enumerate(rows, start=2):
        if row[0] and str(row[1]) in ("", "Count", "Errors", "Value"):
            if str(row[0]).isupper():
                ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)


def _sheet_journey_analysis(
    wb: "Workbook",
    journeys: list[Journey],
    narratives: dict[str, dict[str, str]],
) -> None:
    ws = wb.create_sheet("Journey Analysis")

    headers = [
        "Intent", "Shape", "Health", "Purpose", "What the journey does",
        "Flows", "Pages", "Depth", "Phrases", "Asks for", "Handoff queues",
        "End states", "Coverage gaps", "Recommendation",
    ]

    rows = []
    for j in journeys:
        n = narratives.get(j.intent, {})
        rows.append([
            j.intent,
            _journey_shape(j),
            n.get("health", ""),
            n.get("purpose", ""),
            n.get("what_the_journey_does", ""),
            " · ".join(j.flows),
            j.page_count,
            j.max_depth,
            j.training_phrase_count,
            " · ".join(j.questions_asked),
            " · ".join(j.handoff_queues),
            " · ".join(j.end_states),
            n.get("coverage_gaps", ""),
            n.get("recommendation", ""),
        ])

    rows.sort(key=lambda r: (_HEALTH_ORDER.get(str(r[2]), 3), str(r[0])))
    _write(ws, headers, rows, widths=[30, 15, 16, 46, 62, 26, 8, 8, 10, 24, 30, 30, 46, 52])


def _sheet_response_vs_scope(
    wb: "Workbook",
    journeys: list[Journey],
    narratives: dict[str, dict[str, str]],
) -> None:
    ws = wb.create_sheet("Response vs Scope")

    headers = [
        "Intent", "Verdict", "Purpose", "First response", "All responses",
        "Response critique", "Recommendation",
    ]

    rows = []
    for j in journeys:
        n = narratives.get(j.intent, {})
        responses = j.all_responses
        rows.append([
            j.intent,
            n.get("response_verdict", ""),
            n.get("purpose", ""),
            responses[0] if responses else "",
            _join_capped(responses),
            n.get("response_critique", ""),
            n.get("recommendation", ""),
        ])

    rows.sort(key=lambda r: (_VERDICT_ORDER.get(str(r[1]), 6), str(r[0])))
    _write(ws, headers, rows, widths=[30, 14, 46, 56, 70, 62, 52])


def _sheet_coverage_gaps(
    wb: "Workbook",
    agent: AgentIndex,
    journeys: list[Journey],
    results: list[tuple[str, list[Finding]]],
) -> None:
    ws = wb.create_sheet("Coverage Gaps")

    headers = ["Category", "Intent / Item", "Status", "Detail", "Action"]
    rows: list[list[Any]] = []

    for j in journeys:
        if not j.is_routed:
            rows.append([
                "Unrouted head intent", j.intent, "unrouted",
                f"{j.training_phrase_count} training phrases match an intent no route handles.",
                "Route it, or retire the intent.",
            ])

    for j in journeys:
        if j.is_routed and j.steps and not j.handoff_queues:
            rows.append([
                "No live-agent path", j.intent, "warning",
                f"{j.page_count} page(s), no category assigned anywhere in the journey.",
                "Confirm self-service only is intended.",
            ])

    # Route-group coverage and symmetry come from the routing check, which has
    # already resolved which groups exist and who is missing from them.
    for label, findings in results:
        if label != "Routing & Reachability":
            continue
        for f in findings:
            if "not routed by any head-intent route group" in f.message:
                rows.append([
                    "Missing from route group", _quoted(f.message), "error",
                    f.message, "Add to the head-intent route group.",
                ])
            elif "but not by" in f.message:
                rows.append([
                    "Route group asymmetry", _quoted(f.message), "warning",
                    f.message, "Confirm the asymmetry is deliberate.",
                ])
            elif "has no test case covering it" in f.message:
                rows.append([
                    "No test coverage", _quoted(f.message), "warning",
                    f.message, "Add a test case.",
                ])

    for label, findings in results:
        if label != "Agent Config Integrity":
            continue
        for f in findings:
            if "has no test case covering it" in f.message:
                rows.append([
                    "No test coverage", _quoted(f.message), "warning",
                    f.message, "Add a test case for this head intent.",
                ])

    order = {"unrouted": 0, "error": 1, "warning": 2}
    rows.sort(key=lambda r: (order.get(str(r[2]), 3), str(r[0]), str(r[1])))
    _write(ws, headers, rows, widths=[28, 34, 14, 74, 40])


def _sheet_overlap(wb: "Workbook", agent: AgentIndex) -> None:
    ws = wb.create_sheet("Overlap & Confusion")

    headers = ["Intent A", "Intent B", "Type", "Count", "Score", "Example", "Recommendation"]
    rows: list[list[Any]] = []

    for collision in exact_collisions(agent):
        primary = collision.intents[0]
        for other in collision.intents[1:]:
            rows.append([
                primary, other, "exact", 1, 1.0, collision.phrase,
                "Assign the phrase to one intent and delete the others.",
            ])

    for pair in near_duplicate_pairs(agent):
        rows.append([
            pair.intent_a, pair.intent_b, "near-duplicate", pair.pair_count,
            round(pair.top_score, 2),
            f"\"{pair.example_a}\"  vs  \"{pair.example_b}\"",
            "Rule on the boundary between these intents.",
        ])

    rows.sort(key=lambda r: (0 if r[2] == "exact" else 1, -int(r[3]), str(r[0])))
    _write(ws, headers, rows, widths=[30, 30, 16, 9, 9, 76, 46])


def _sheet_asr(wb: "Workbook", agent: AgentIndex) -> None:
    ws = wb.create_sheet("ASR & Noise Phrases")

    headers = ["Intent", "Token", "Likely word", "Phrases", "Example", "Action"]
    rows = [
        [
            hit.intent, hit.token, hit.likely_word, hit.phrase_count,
            hit.example,
            f"Confirm '{hit.token}' is genuine, not a mistranscription of "
            f"'{hit.likely_word}'.",
        ]
        for hit in asr_hits(agent)
    ]

    rows.sort(key=lambda r: (-int(r[3]), str(r[0])))
    _write(ws, headers, rows, widths=[30, 14, 14, 10, 74, 56])


def _sheet_data_hygiene(wb: "Workbook", agent: AgentIndex) -> None:
    ws = wb.create_sheet("Data Hygiene")

    headers = ["Issue type", "Intent", "Severity", "Detail", "Action"]
    rows: list[list[Any]] = []

    for stat in phrase_stats(agent):
        if stat.empty:
            rows.append([
                "Empty training phrases", stat.intent, "error",
                f"{stat.empty} phrase(s) are empty strings.",
                "Delete them.",
            ])
        if stat.within_duplicates:
            rows.append([
                "Within-intent duplicates", stat.intent, "warning",
                f"{stat.within_duplicates} phrase(s) appear more than once "
                f"({stat.total} total, {stat.unique} unique).",
                "Deduplicate.",
            ])
        if stat.drifted:
            rows.append([
                "Metadata drift", stat.intent, "warning",
                f"Declares {stat.declared} phrases, {stat.total} exported.",
                "Re-export or correct numTrainingPhrases.",
            ])
        if stat.is_head and stat.total < 10:
            rows.append([
                "Thin head intent", stat.intent, "warning",
                f"Only {stat.total} training phrase(s) for a head intent.",
                "Add phrases or merge into a sibling intent.",
            ])

    order = {"error": 0, "warning": 1}
    rows.sort(key=lambda r: (order.get(str(r[2]), 2), str(r[0]), str(r[1])))
    _write(ws, headers, rows, widths=[28, 34, 12, 62, 44])


def _sheet_phrase_counts(wb: "Workbook", agent: AgentIndex) -> None:
    ws = wb.create_sheet("Phrase Counts")

    headers = ["Intent", "Kind", "Phrases", "Unique", "Duplicates", "Health"]
    rows = [
        [
            stat.intent,
            "head" if stat.is_head else (", ".join(stat.labels) or "unlabelled"),
            stat.total, stat.unique, stat.within_duplicates, stat.health,
        ]
        for stat in phrase_stats(agent)
    ]

    rows.sort(key=lambda r: (_HEALTH_ORDER.get(str(r[5]), 3), int(r[2])))
    _write(ws, headers, rows, widths=[34, 22, 11, 11, 13, 18])


def _sheet_findings(wb: "Workbook", results: list[tuple[str, list[Finding]]]) -> None:
    ws = wb.create_sheet("Validation Findings")

    headers = ["Check", "Severity", "File", "Flow", "Message", "Detail"]
    rows = [
        [label, f.severity, f.file_path, f.flow_name, f.message, f.detail]
        for label, findings in results
        for f in findings
        if f.severity != "pass"  # Passes belong in the HTML report, not here
    ]

    order = {"error": 0, "warning": 1}
    rows.sort(key=lambda r: (order.get(str(r[1]), 2), str(r[0]), str(r[2])))
    _write(ws, headers, rows, widths=[28, 12, 52, 22, 78, 66])


# ── Shared writer ─────────────────────────────────────────────────────────────

def _write(
    ws: "Worksheet",
    headers: list[str],
    rows: Iterable[list[Any]],
    widths: list[int],
) -> None:
    """Write a header row plus data, then apply consistent formatting."""
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    header_font = Font(bold=True, color=_HEADER_FONT, size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    top = Alignment(vertical="top")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    coloured_cols = {
        idx for idx, name in enumerate(headers, start=1)
        if name in _COLOURED_HEADERS
    }

    row_count = 0
    for row in rows:
        ws.append(row)
        row_count += 1
        excel_row = row_count + 1
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            # Long free-text columns wrap; short ones stay on one line so the
            # sheet does not become mostly whitespace.
            cell.alignment = wrap if widths[col_idx - 1] >= 40 else top
            if col_idx in coloured_cols:
                colour = _VALUE_COLOURS.get(str(cell.value).strip().lower())
                if colour:
                    cell.fill = PatternFill("solid", fgColor=colour)

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    if row_count:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_count + 1}"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _journey_shape(journey: Journey) -> str:
    if not journey.is_routed:
        return "unrouted"
    if journey.answered_inline:
        return "inline answer"
    if not journey.steps:
        return "routed, no response"
    return "page flow"


def _pct(part: int, whole: int) -> str:
    return f"{part / whole:.0%}" if whole else ""


def _join_capped(values: list[str], limit: int = 30) -> str:
    """Join responses, capping length so a cell stays inside Excel's limit."""
    joined = "  |  ".join(values[:limit])
    if len(values) > limit:
        joined += f"  |  (+{len(values) - limit} more)"
    # Excel rejects cell values over 32,767 characters.
    return joined[:32000]


def _quoted(message: str) -> str:
    """Pull the first single-quoted name out of a finding message."""
    start = message.find("'")
    if start == -1:
        return message[:60]
    end = message.find("'", start + 1)
    return message[start + 1:end] if end != -1 else message[:60]
